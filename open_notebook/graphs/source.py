import asyncio
import json
import operator
import os
import re
import shutil
from typing import Any, Awaitable, Callable, Dict, List, Optional, TypeVar

from content_core import extract_content
from content_core.common import ProcessSourceState
from langchain_core.runnables import RunnableConfig
from langgraph.graph import END, START, StateGraph
from langgraph.types import Send
from loguru import logger
from pydantic import BaseModel, Field
from typing_extensions import Annotated, TypedDict

from open_notebook.ai.models import Model, ModelManager
from open_notebook.domain.content_settings import ContentSettings
from open_notebook.domain.notebook import Asset, Source
from open_notebook.domain.transformation import Transformation
from open_notebook.graphs.transformation import graph as transform_graph
from open_notebook.utils.text_utils import clean_thinking_content

T = TypeVar("T")

SUPPORTED_VISION_IMAGE_EXTENSIONS = (
    ".png",
    ".jpg",
    ".jpeg",
    ".gif",
    ".webp",
    ".bmp",
    ".tif",
    ".tiff",
    ".img",
)


def should_bypass_content_core_for_image(file_path: str | None) -> bool:
    if not file_path:
        return False
    return os.path.splitext(file_path)[1].lower() in SUPPORTED_VISION_IMAGE_EXTENSIONS


class SourceState(TypedDict):
    content_state: ProcessSourceState
    apply_transformations: List[Transformation]
    source_id: str
    notebook_ids: List[str]
    source: Source
    transformation: Annotated[list, operator.add]
    embed: bool
    language: Optional[str]


class FigureContext(BaseModel):
    filename: str
    source_kind: str = "unknown"
    page: Optional[int] = None
    sheet_name: Optional[str] = None
    cell_anchor: Optional[str] = None
    image_role: Optional[str] = None
    nearby_text: str = ""
    table_headers: List[str] = Field(default_factory=list)
    table_row_text: str = ""
    image_url: str = ""
    width: Optional[int] = None
    height: Optional[int] = None


class VisionDescription(BaseModel):
    image_type: str = "unknown"
    readability: str = "unknown"
    confidence: float = 0.0
    description_level: str = "full"
    confirmed_facts: List[str] = Field(default_factory=list)
    extracted_values: List[Dict[str, Any]] = Field(default_factory=list)
    uncertain_items: List[str] = Field(default_factory=list)
    domain_interpretation: str = ""


class VisionDescriptionResult(BaseModel):
    accepted: bool
    description: Optional[VisionDescription] = None
    rejection_reasons: List[str] = Field(default_factory=list)
    raw_text: str = ""


def _env_int(name: str, default: int) -> int:
    raw = os.environ.get(name)
    if raw is None or raw.strip() == "":
        return default
    try:
        return int(raw)
    except ValueError:
        logger.warning(f"Invalid integer for {name}={raw!r}; using default {default}")
        return default


def _env_float(name: str, default: float) -> float:
    raw = os.environ.get(name)
    if raw is None or raw.strip() == "":
        return default
    try:
        return float(raw)
    except ValueError:
        logger.warning(f"Invalid float for {name}={raw!r}; using default {default}")
        return default


def _vision_concurrency() -> int:
    return max(1, _env_int("VISION_CONCURRENCY", 6))


def _is_transient_vision_error(error: BaseException) -> bool:
    """Return whether a vision call error is worth retrying."""
    if isinstance(error, TimeoutError):
        return True

    message = str(error).lower()
    transient_markers = (
        "429",
        "500",
        "502",
        "503",
        "504",
        "520",
        "rate_limit",
        "rate limit",
        "too many requests",
        "server_error",
        "temporarily",
        "timeout",
        "timed out",
        "connection reset",
        "connection error",
        "remote protocol error",
    )
    return any(marker in message for marker in transient_markers)


async def _invoke_vision_with_retries(call: Callable[[], Awaitable[T]]) -> T:
    """Invoke a single vision request with timeout and bounded transient retries."""
    timeout_seconds = _env_float("VISION_TIMEOUT_SECONDS", 120.0)
    max_retries = max(0, _env_int("VISION_MAX_RETRIES", 2))
    base_delay_seconds = max(0.0, _env_float("VISION_RETRY_BASE_DELAY_SECONDS", 3.0))

    last_error: BaseException | None = None
    for attempt in range(max_retries + 1):
        try:
            return await asyncio.wait_for(call(), timeout=timeout_seconds)
        except Exception as error:
            last_error = error
            if attempt >= max_retries or not _is_transient_vision_error(error):
                raise
            delay = base_delay_seconds * (2**attempt)
            logger.warning(
                f"Vision request failed transiently; retrying "
                f"{attempt + 1}/{max_retries} after {delay:.1f}s: {error}"
            )
            if delay > 0:
                await asyncio.sleep(delay)

    raise RuntimeError("Vision request retry loop exited unexpectedly") from last_error


def _vision_failure_description(error: BaseException) -> str:
    """User-facing fallback when a remote vision request ultimately fails."""
    if isinstance(error, TimeoutError) or "timeout" in str(error).lower():
        reason = "图片解析服务响应超时"
    elif _is_transient_vision_error(error):
        reason = "图片解析服务临时繁忙"
    else:
        reason = "图片解析服务未能完成该图描述"

    return (
        f"{reason}，已保留原图，建议稍后重新解析该图片。"
    )


def _vision_model_inference_kwargs(provider: str) -> Dict[str, Any]:
    """Return provider-specific inference options for document image descriptions."""
    normalized_provider = provider.replace("_", "-").lower()
    temperature = _env_float("VISION_TEMPERATURE", 0.0)

    if normalized_provider == "ollama":
        return {
            "num_ctx": _env_int("VISION_NUM_CTX", 2048),
            "num_predict": _env_int("VISION_NUM_PREDICT", 384),
            "temperature": temperature,
        }

    return {
        "max_tokens": _env_int("VISION_MAX_TOKENS", 384),
        "temperature": temperature,
    }


VISION_PROMPT = """You are a senior oilfield chemist specializing in oil well cement admixture analysis.
Analyze figures from technical reports for fluid loss additives, retarders, dispersants,
anti-channeling agents, flushing agents, weighting materials, and plugging particles.

Classify the image using one of these values:
- hpht_curve: HPHT consistometer chart with temperature, pressure, consistency/Bc over time.
- analytical_spectrum: FTIR, GC-MS, HPLC, XRD, SEM/TEM or related analytical result.
- lab_photo: physical cement slurry, cured sample, mixing tool, slurry cup, mold, or test artifact.
- performance_comparison: plots comparing dosage, fluid loss, rheology, compressive strength, free water, or stability.
- mechanism_schematic: synthesis process, molecular structure, adsorption or crosslinking mechanism.
- embedded_table_or_screenshot: spreadsheet/table screenshot, composite report block, or chart embedded in a table.
- unknown: image does not fit the above or is unreadable.

Domain vocabulary:
- 稠度(Bc) = Consistency in Bearden units.
- 稠化时间 = Thickening time.
- 温度 = Temperature.
- 压力 = Pressure.
- 流变 = Rheology, commonly six-speed viscometer readings.
- 包芯 = agglomeration or wrapped core after test.
- 沉降/沉死 = sedimentation or hard settling.
- 自由水 = free water.
- 养护 = curing.

Use the supplied document context as supporting evidence. Do not invent values that are
not visible in the image or present in the context. If a curve color cannot be mapped to
temperature, pressure, or consistency with confidence, put that limitation in uncertain_items.

Return only one JSON object with this schema:
{
  "image_type": "hpht_curve|analytical_spectrum|lab_photo|performance_comparison|mechanism_schematic|embedded_table_or_screenshot|unknown",
  "readability": "high|medium|low|unreadable",
  "confidence": 0.0,
  "description_level": "full|trend_only|visual_state_only|context_only|reject",
  "confirmed_facts": ["facts visible in image or supported by context"],
  "extracted_values": [{"name": "parameter", "value": "value with unit when available", "evidence": "visible/context"}],
  "uncertain_items": ["items that cannot be confirmed"],
  "domain_interpretation": "short professional interpretation tied to cement admixture R&D"
}

For low-resolution charts, still return JSON. Use description_level="trend_only",
do not extract exact numeric values, and describe only visible curve trends.
For physical sample photos, use description_level="visual_state_only" and describe
visible slurry/sample/tool state. For table screenshots where context contains the
row text, use description_level="context_only" and cite the row text as evidence."""


def _build_vision_prompt(
    language: Optional[str],
    figure_context: Optional[FigureContext] = None,
) -> str:
    """Return the vision model prompt with language instruction at top."""
    if language and language.startswith("en"):
        lang_instruction = "IMPORTANT: You MUST respond in English. All output must be in English.\n\n"
    else:
        lang_instruction = "重要：你必须使用简体中文回复。所有输出必须为中文。\n\n"
    context_text = ""
    if figure_context:
        context_items = {
            "filename": figure_context.filename,
            "source_kind": figure_context.source_kind,
            "page": figure_context.page,
            "sheet_name": figure_context.sheet_name,
            "cell_anchor": figure_context.cell_anchor,
            "image_role": figure_context.image_role,
            "table_headers": figure_context.table_headers,
            "table_row_text": figure_context.table_row_text,
            "nearby_text": figure_context.nearby_text,
        }
        context_text = (
            "\n\nDocument context for this image:\n"
            + json.dumps(context_items, ensure_ascii=False)
        )
    return lang_instruction + VISION_PROMPT + context_text


_REASONING_PREFIX_PATTERNS = [
    r"^用户希望我.*?\n",
    r"^用户要求我.*?\n",
    r"^我需要.*?\n",
    r"^让我.*?\n",
    r"^首先[，,].*?\n",
    r"^我们.*?查看.*?\n",
    r"^等等[，,].*?\n",
    r"^重新审视.*?\n",
    r"^让我们.*?\n",
    r"^看这.*?\n",
    r"^如果[^，,\n]*[，,].*?\n",
    r"^通常[^，,\n]*[，,].*?\n",
    r"^但是[，,].*?\n",
    r"^不[，,].*?\n",
    r"^修正[：:].*?\n",
    r"^观察图片.*?\n",
    r"^观察[：:].*?\n",
    r"^假设[^，,\n]*[，,].*?\n",
    r"^或者[，,].*?\n",
    r"^再看.*?\n",
    r"^仔细看.*?\n",
    r"^可能[^，,\n]*[，,].*?\n",
    r"^也许[^，,\n]*[，,].*?\n",
    r"^但是[^，,\n]*如果.*?\n",
    r"^分析步骤[：:].*?\n",
    r"^根据分类.*?\n",
    r"^首先[，,]我需要.*?\n",
    r"^现在.*?分析.*?\n",
    r"^接下来.*?\n",
]

_STRUCTURED_HEADER_RE = re.compile(
    r"^[-*]?\s*("
    r"Image type|图像类型|图片类型|图片内容|分类"
    r"|Domain analysis|领域分析"
    r"|Key data|关键数据"
    r"|Quality observations|质量观察"
    r"|Figure type|图类型|Figure"
    r")",
    re.IGNORECASE | re.MULTILINE,
)


_PROMPT_LEAKAGE_MARKERS = [
    "CRITICAL RULES",
    "VIOLATION MEANS FAILURE",
    "必须遵守的规则",
    "Output structure",
    "Do NOT echo",
    "NEVER output",
    "规则原文",
]

_REASONING_LEAKAGE_MARKERS = [
    "让我们",
    "假设",
    "修正",
    "重新审视",
    "等等",
    "仔细看",
    "我需要",
    "接下来",
]


def _extract_json_object(raw: str) -> Optional[Dict[str, Any]]:
    text = clean_thinking_content(raw).strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE)
        text = re.sub(r"\s*```$", "", text)

    start = text.find("{")
    end = text.rfind("}")
    if start < 0 or end <= start:
        return None

    try:
        parsed = json.loads(text[start : end + 1])
    except json.JSONDecodeError:
        return None
    return parsed if isinstance(parsed, dict) else None


def _contains_any_marker(text: str, markers: List[str]) -> bool:
    return any(marker in text for marker in markers)


def _structured_vision_description(raw: str) -> VisionDescriptionResult:
    parsed = _extract_json_object(raw)
    if parsed is None:
        return VisionDescriptionResult(
            accepted=False,
            rejection_reasons=["invalid_json"],
            raw_text=raw,
        )

    try:
        description = VisionDescription.model_validate(parsed)
    except Exception:
        return VisionDescriptionResult(
            accepted=False,
            rejection_reasons=["invalid_schema"],
            raw_text=raw,
        )

    reasons: List[str] = []
    combined = json.dumps(parsed, ensure_ascii=False)
    if _contains_any_marker(combined, _PROMPT_LEAKAGE_MARKERS):
        reasons.append("prompt_leakage")
    if _contains_any_marker(combined, _REASONING_LEAKAGE_MARKERS):
        reasons.append("reasoning_leakage")
    if description.confidence < 0.35:
        description.uncertain_items.append("模型对该图片的置信度较低，描述仅作保守参考。")
    if description.readability in ("low", "unreadable") and description.extracted_values:
        description.extracted_values = []
        description.uncertain_items.append("图片清晰度不足，已忽略模型给出的具体定量值。")

    return VisionDescriptionResult(
        accepted=not reasons,
        description=description,
        rejection_reasons=reasons,
        raw_text=raw,
    )


def _extract_partial_json_string(raw: str, key: str) -> str:
    match = re.search(rf'"{re.escape(key)}"\s*:\s*"([^"]*)"', raw)
    return match.group(1).strip() if match else ""


def _extract_partial_json_facts(raw: str) -> List[str]:
    facts: List[str] = []
    for value in re.findall(r'"([^"]+)"', raw):
        stripped = value.strip()
        if not stripped:
            continue
        if re.fullmatch(r"[a-zA-Z_]+", stripped):
            continue
        if stripped in facts:
            continue
        facts.append(stripped)
    return facts[:6]


def _partial_json_vision_description(raw: str) -> Optional[VisionDescriptionResult]:
    image_type = _extract_partial_json_string(raw, "image_type")
    valid_types = {
        "hpht_curve",
        "analytical_spectrum",
        "lab_photo",
        "performance_comparison",
        "mechanism_schematic",
        "embedded_table_or_screenshot",
        "unknown",
    }
    if image_type not in valid_types:
        return None

    readability = _extract_partial_json_string(raw, "readability") or "low"
    description_level = _extract_partial_json_string(raw, "description_level") or "context_only"
    facts = _extract_partial_json_facts(raw)
    facts = [fact for fact in facts if fact not in valid_types and fact != readability and fact != description_level]
    if not facts:
        return None

    return VisionDescriptionResult(
        accepted=True,
        description=VisionDescription(
            image_type=image_type,
            readability=readability,
            confidence=0.55,
            description_level=description_level,
            confirmed_facts=facts,
            extracted_values=[],
            uncertain_items=["模型返回格式不完整，已保留可读事实并忽略不可靠定量值。"],
            domain_interpretation="该描述来自不完整结构化结果，需结合原图复核。",
        ),
        rejection_reasons=[],
        raw_text=raw,
    )


def _infer_fallback_image_type(text: str, context: FigureContext) -> str:
    image_text = text.lower()
    combined = f"{image_text}\n{context.table_row_text}\n{context.nearby_text}".lower()
    explicit_type = _extract_partial_json_string(text, "image_type")
    if explicit_type in {
        "hpht_curve",
        "analytical_spectrum",
        "lab_photo",
        "performance_comparison",
        "mechanism_schematic",
        "embedded_table_or_screenshot",
        "unknown",
    }:
        return explicit_type
    if any(keyword in combined for keyword in ["表格", "截图", "流变", "稳定性", "加量"]):
        return "embedded_table_or_screenshot"
    photo_keywords = ["照片", "浆杯", "搅拌", "叶片", "水泥浆", "浆体"]
    if any(keyword in image_text for keyword in photo_keywords):
        return "lab_photo"
    if any(keyword in combined for keyword in ["hpht", "稠化", "稠度", "温度", "压力", "曲线"]):
        return "hpht_curve"
    if any(keyword in combined for keyword in ["包芯", "沉降", "沉死"]):
        return "lab_photo"
    if any(keyword in combined for keyword in ["ftir", "红外", "sem", "谱", "粒径", "电镜"]):
        return "analytical_spectrum"
    return "unknown"


def _fallback_description_level(image_type: str, text: str, context: FigureContext) -> str:
    combined = f"{text}\n{context.table_row_text}\n{context.nearby_text}"
    if image_type == "hpht_curve":
        return "trend_only"
    if image_type == "lab_photo":
        return "visual_state_only"
    if context.table_row_text.strip() or image_type == "embedded_table_or_screenshot":
        return "context_only"
    return "context_only"


def _clean_fallback_text(raw: str) -> str:
    text = clean_thinking_content(raw).strip()
    text = re.sub(r"```(?:json)?", "", text, flags=re.IGNORECASE).replace("```", "")
    for pattern in _REASONING_PREFIX_PATTERNS:
        text = re.sub(pattern, "", text, count=10)
    lines = [line.strip() for line in text.splitlines()]
    lines = [
        line
        for line in lines
        if line
        and line not in ("{", "}", "[", "]")
        and not re.match(r'^"?[a-zA-Z_]+"?\s*:', line)
    ]
    return "\n".join(lines).strip()


def _fallback_facts_from_text(text: str) -> List[str]:
    facts: List[str] = []
    for line in text.splitlines():
        cleaned = re.sub(r"^[-*]\s*", "", line).strip()
        cleaned = cleaned.rstrip(",")
        if cleaned in ("{", "}", "[", "]"):
            continue
        if re.match(r'^"?[a-zA-Z_]+"?\s*:', cleaned):
            continue
        quoted = re.fullmatch(r'"([^"]+)"', cleaned)
        if quoted:
            cleaned = quoted.group(1).strip()
        cleaned = re.sub(r"^(图像类型|可确认信息|无法确认|领域解释)[：:].*$", "", cleaned).strip()
        if cleaned and cleaned not in facts:
            facts.append(cleaned)
    if not facts and text:
        facts.append(text)
    return facts[:6]


def _has_fallback_signal(text: str, context: FigureContext) -> bool:
    combined = f"{text}\n{context.table_row_text}\n{context.nearby_text}"
    signals = [
        "图像类型",
        "曲线",
        "趋势",
        "温度",
        "压力",
        "稠度",
        "照片",
        "浆杯",
        "搅拌",
        "叶片",
        "水泥浆",
        "浆体",
        "包芯",
        "沉降",
        "沉死",
        "流变",
        "稳定性",
        "表格",
        "截图",
    ]
    return any(signal in combined for signal in signals)


def _fallback_vision_description(
    raw: str,
    context: FigureContext,
    rejection_reasons: List[str],
) -> VisionDescriptionResult:
    combined_raw = clean_thinking_content(raw)
    if _contains_any_marker(combined_raw, _PROMPT_LEAKAGE_MARKERS):
        return _minimum_vision_description(
            context,
            ["prompt_leakage"],
        )
    if _contains_any_marker(combined_raw, _REASONING_LEAKAGE_MARKERS):
        return _minimum_vision_description(
            context,
            ["reasoning_leakage"],
        )

    partial_result = _partial_json_vision_description(combined_raw)
    if partial_result is not None:
        return partial_result

    text = _clean_fallback_text(raw)
    if not text or not _has_fallback_signal(text, context):
        return _minimum_vision_description(context, rejection_reasons)

    image_type = _infer_fallback_image_type(text, context)
    description_level = _fallback_description_level(image_type, text, context)
    confirmed_facts = _fallback_facts_from_text(text)
    if context.table_row_text.strip():
        confirmed_facts.append(f"相邻表格行记录：{context.table_row_text.strip()}")

    uncertain_items = []
    if description_level == "trend_only":
        uncertain_items.append("图片清晰度不足，具体坐标值、曲线颜色含义和终点时间需结合原始图或表格复核。")
    elif description_level == "visual_state_only":
        uncertain_items.append("仅能描述可见实物状态，配方条件和定量指标需结合相邻表格复核。")
    elif description_level == "context_only":
        uncertain_items.append("描述主要依赖图片附近的表格或正文，图中细节需人工复核。")

    return VisionDescriptionResult(
        accepted=True,
        description=VisionDescription(
            image_type=image_type,
            readability="low" if description_level == "trend_only" else "medium",
            confidence=0.55,
            description_level=description_level,
            confirmed_facts=confirmed_facts,
            extracted_values=[],
            uncertain_items=uncertain_items,
            domain_interpretation="该描述为质量降级后的保守说明，仅保留可见趋势、实物状态或相邻表格证据。",
        ),
        rejection_reasons=[],
        raw_text=raw,
    )


def _minimum_vision_description(
    context: FigureContext,
    reasons: Optional[List[str]] = None,
) -> VisionDescriptionResult:
    if context.width and context.height and (context.width < 160 or context.height < 120):
        intro = "图片较小，以下描述可信度不高；已保留原图供用户查看。"
    elif context.width and context.height:
        intro = "模型未能稳定识别该图片内容，已保留原图供用户查看。"
    else:
        intro = "图片内容未能稳定识别，以下描述可信度不高；已保留原图供用户查看。"
    facts = [intro]
    if context.table_row_text.strip():
        facts.append(f"相邻表格行记录：{context.table_row_text.strip()}")
    if context.width and context.height:
        facts.append(f"图片尺寸：{context.width}x{context.height}px")

    return VisionDescriptionResult(
        accepted=True,
        description=VisionDescription(
            image_type="unknown",
            readability="low",
            confidence=0.0,
            description_level="context_only",
            confirmed_facts=facts,
            extracted_values=[],
            uncertain_items=["无法从当前图片中可靠提取具体数据或专业结论。"],
            domain_interpretation="该图片需要结合原图、相邻正文或表格进行人工复核。",
        ),
        rejection_reasons=[],
        raw_text="",
    )


def _format_context_line(context: FigureContext) -> str:
    parts = []
    if context.page is not None:
        parts.append(f"页码：{context.page}")
    if context.cell_anchor:
        parts.append(f"位置：{context.cell_anchor}")
    elif context.sheet_name:
        parts.append(f"工作表：{context.sheet_name}")
    return "；".join(parts)


def _render_vision_result(
    filename: str,
    result: VisionDescriptionResult,
    context: FigureContext,
) -> str:
    if not result.accepted or result.description is None:
        result = _minimum_vision_description(context, result.rejection_reasons)

    desc = result.description
    if desc is None:
        desc = _minimum_vision_description(context).description
    if desc is None:
        return "图像类型：unknown\n描述级别：unknown\n可确认信息：\n- 图片内容无法可靠识别，保留原图供用户查看。"
    lines = [
        f"图像类型：{desc.image_type}",
        f"可读性：{desc.readability}",
        f"置信度：{desc.confidence:.2f}",
        f"描述级别：{desc.description_level}",
    ]

    context_line = _format_context_line(context)
    if context_line:
        lines.append(f"来源位置：{context_line}")
    if context.table_headers:
        lines.append("关联表头：" + " | ".join(str(v) for v in context.table_headers if str(v).strip()))
    if context.table_row_text:
        lines.append(f"关联表格行：{context.table_row_text}")

    if desc.confirmed_facts:
        lines.append("可确认信息：")
        lines.extend(f"- {fact}" for fact in desc.confirmed_facts if fact.strip())

    if desc.extracted_values:
        lines.append("提取数据：")
        for item in desc.extracted_values:
            name = str(item.get("name", "")).strip()
            value = str(item.get("value", "")).strip()
            evidence = str(item.get("evidence", "")).strip()
            if not name and not value:
                continue
            suffix = f"（依据：{evidence}）" if evidence else ""
            lines.append(f"- {name}：{value}{suffix}")

    if desc.uncertain_items:
        lines.append("无法确认：")
        lines.extend(f"- {item}" for item in desc.uncertain_items if item.strip())

    if desc.domain_interpretation.strip():
        lines.append(f"领域解释：{desc.domain_interpretation.strip()}")

    return "\n".join(lines)


def _safe_figure_description(
    filename: str,
    raw: str,
    context: FigureContext,
) -> str:
    result = _structured_vision_description(raw)
    if not result.accepted and "invalid_json" in result.rejection_reasons:
        result = _fallback_vision_description(raw, context, result.rejection_reasons)
    return _render_vision_result(filename, result, context)


def _clean_vision_response(raw: str) -> str:
    """Strip reasoning artifacts and extract structured output from vision model response."""
    text = clean_thinking_content(raw)

    # If the response starts with a structured header, return as-is
    if _STRUCTURED_HEADER_RE.match(text.strip()):
        return text.strip()

    # Remove reasoning prefix lines (internal monologue)
    for pattern in _REASONING_PREFIX_PATTERNS:
        text = re.sub(pattern, "", text, count=10)

    # Strip leading blank lines from cleanup
    text = text.strip()

    # If structured header exists after cleanup, trim everything before the FIRST one
    match = _STRUCTURED_HEADER_RE.search(text)
    if match:
        text = text[match.start():]
        return text.strip()

    # Fallback: no structured header at all — keep only the last 30% (conclusions)
    lines = text.split("\n")
    if len(lines) > 6:
        keep_start = max(0, int(len(lines) * 0.7))
        text = "\n".join(lines[keep_start:])

    return text.strip()


class TransformationState(TypedDict):
    source: Source
    transformation: Transformation


def _sanitize_excel_table_newlines(content: str) -> str:
    """Fix broken markdown table rows caused by newlines in Excel cells.

    content_core's openpyxl-based Excel extraction inserts raw \\n into
    markdown table rows, which splits a single row into multiple lines.
    This merges orphan continuation lines back into the preceding table row
    using <br> to preserve multi-line semantics.
    """
    def _is_table_separator_line(line: str) -> bool:
        stripped = line.strip()
        if not stripped.startswith("|"):
            return False
        body = stripped.replace("|", "").replace("-", "").replace(":", "").replace(" ", "")
        return body == ""

    lines = content.split("\n")
    result = []
    expected_pipes = None
    i = 0

    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        if not stripped.startswith("|"):
            expected_pipes = None
            result.append(line)
            i += 1
            continue

        if _is_table_separator_line(line):
            expected_pipes = line.count("|")
            result.append(line)
            i += 1
            continue

        if expected_pipes is None:
            expected_pipes = line.count("|")

        merged = line
        while merged.count("|") < expected_pipes and i + 1 < len(lines):
            next_line = lines[i + 1]
            next_stripped = next_line.strip()

            if not next_stripped or next_stripped.startswith("#"):
                break

            merged += "<br>" + next_line
            i += 1

        result.append(merged)
        i += 1

    return "\n".join(result)


def _trim_excel_empty_table_rows(content: str) -> str:
    """Remove fully empty Markdown table rows emitted from formatted Excel ranges."""
    result: List[str] = []
    for line in content.splitlines():
        stripped = line.strip()
        if stripped.startswith("|") and stripped.endswith("|"):
            cells = stripped[1:-1].split("|")
            if cells and all(not cell.strip() for cell in cells):
                continue
        result.append(line)
    return "\n".join(result)


def _excel_col_name(index_1_based: int) -> str:
    name = ""
    index = max(index_1_based, 1)
    while index:
        index, rem = divmod(index - 1, 26)
        name = chr(65 + rem) + name
    return name


def _infer_excel_image_ext(image_obj: Any) -> str:
    path = getattr(image_obj, "path", "") or ""
    ext = os.path.splitext(path)[1].lower().lstrip(".")
    if ext:
        return ext
    fmt = (getattr(image_obj, "format", "") or "").lower()
    if fmt:
        return fmt
    return "png"


def _extract_excel_image_bytes(image_obj: Any) -> Optional[bytes]:
    data_fn = getattr(image_obj, "_data", None)
    if callable(data_fn):
        try:
            data = data_fn()
            if isinstance(data, (bytes, bytearray)):
                return bytes(data)
        except Exception:
            return None
    return None


def _excel_anchor_label(image_obj: Any, sheet_title: str) -> str:
    anchor = getattr(image_obj, "anchor", None)
    anchor_from = getattr(anchor, "_from", None)
    if anchor_from is None:
        return ""
    try:
        col = int(getattr(anchor_from, "col", 0)) + 1
        row = int(getattr(anchor_from, "row", 0)) + 1
    except Exception:
        return ""
    return f"{sheet_title}!{_excel_col_name(col)}{row}"


def _excel_anchor_cell(anchor_label: str) -> str:
    if "!" in anchor_label:
        return anchor_label.rsplit("!", 1)[-1]
    return anchor_label


def _excel_cell_row_index(anchor_cell: str) -> Optional[int]:
    match = re.search(r"(\d+)$", anchor_cell or "")
    if not match:
        return None
    return max(int(match.group(1)) - 1, 0)


def _stringify_excel_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def _build_excel_context_from_anchor(
    rows: List[List[Any]],
    sheet_title: str,
    anchor_cell: str,
    filename: str,
    width: Optional[int] = None,
    height: Optional[int] = None,
) -> FigureContext:
    row_index = _excel_cell_row_index(anchor_cell)
    table_headers: List[str] = []
    table_row_text = ""

    if row_index is not None and 0 <= row_index < len(rows):
        row = rows[row_index]
        table_row_text = " | ".join(
            value for value in (_stringify_excel_cell(cell) for cell in row) if value
        )

        for header_index in range(row_index - 1, -1, -1):
            header_values = [
                _stringify_excel_cell(cell) for cell in rows[header_index]
            ]
            header_values = [value for value in header_values if value]
            if header_values:
                table_headers = header_values
                break

    full_anchor = f"{sheet_title}!{anchor_cell}" if anchor_cell else sheet_title
    return FigureContext(
        filename=filename,
        source_kind="excel",
        sheet_name=sheet_title,
        cell_anchor=full_anchor,
        image_role="embedded_excel_image",
        table_headers=table_headers,
        table_row_text=table_row_text,
        width=width,
        height=height,
    )


def _build_excel_figure_markdown(
    safe_source_id: str,
    figures: List[Dict[str, str]],
    descriptions_by_file: Dict[str, str],
) -> str:
    figures_section_parts = ["\n\n## Extracted Figures\n"]
    descriptions_section_parts = ["\n\n## Figure Descriptions\n"]

    for fig_index, fig in enumerate(figures, start=1):
        filename = fig["filename"]
        anchor = fig.get("anchor", "")
        anchor_text = f" ({anchor})" if anchor else ""
        figures_section_parts.append(
            f"\n### Figure {fig_index}{anchor_text}\n"
            f"![](/api/uploads/images/{safe_source_id}/{filename})\n"
        )
        descriptions_section_parts.append(
            f"\n### Figure {fig_index}\n"
            f"{descriptions_by_file.get(filename, 'Description unavailable.')}\n"
        )

    return "".join(figures_section_parts) + "".join(descriptions_section_parts)


def _build_standalone_image_markdown(
    safe_source_id: str,
    image_files: List[str],
    descriptions_by_file: Dict[str, str],
) -> str:
    images_section_parts = ["\n\n## Extracted Images\n"]
    descriptions_section_parts = ["\n\n## Figure Descriptions\n"]

    for image_index, filename in enumerate(image_files, start=1):
        images_section_parts.append(
            f"\n### Image {image_index}\n"
            f"![](/api/uploads/images/{safe_source_id}/{filename})\n"
        )
        descriptions_section_parts.append(
            f"\n### Figure: {filename}\n"
            f"{descriptions_by_file.get(filename, 'Description unavailable.')}\n"
        )

    return "".join(images_section_parts) + "".join(descriptions_section_parts)


def _default_figure_context(
    img_file: str,
    is_excel_source: bool,
    safe_source_id: str,
) -> FigureContext:
    if is_excel_source:
        return FigureContext(
            filename=img_file,
            source_kind="excel",
            image_role="embedded_excel_image",
            image_url=f"/api/uploads/images/{safe_source_id}/{img_file}",
        )

    return FigureContext(
        filename=img_file,
        source_kind="image",
        image_role="standalone_image",
        image_url=f"/api/uploads/images/{safe_source_id}/{img_file}",
    )


async def content_process(state: SourceState) -> dict:
    ContentSettings.clear_instance()  # Force reload from DB
    content_settings = await ContentSettings.get_instance()
    content_state: Dict[str, Any] = state["content_state"]  # type: ignore[assignment]

    content_state["url_engine"] = (
        content_settings.default_content_processing_engine_url or "auto"
    )
    content_state["document_engine"] = (
        content_settings.default_content_processing_engine_doc or "auto"
    )
    content_state["output_format"] = "markdown"

    # Add speech-to-text model configuration from Default Models
    try:
        model_manager = ModelManager()
        defaults = await model_manager.get_defaults()
        if defaults.default_speech_to_text_model:
            stt_model = await Model.get(defaults.default_speech_to_text_model)
            if stt_model:
                content_state["audio_provider"] = stt_model.provider
                content_state["audio_model"] = stt_model.name
                logger.info(
                    f"Using speech-to-text model: {stt_model.provider}/{stt_model.name}"
                )
    except Exception as e:
        logger.warning(f"Failed to retrieve speech-to-text model configuration: {e}")
        # Continue without custom audio model (content-core will use its default)

    logger.info(f"Starting content extraction for source_id={state.get('source_id')}")
    logger.info(f"Engine doc: {content_state.get('document_engine')}, URL: {content_state.get('url_engine')}")
    extracted_excel_figures: List[Dict[str, str]] = []
    figure_contexts_by_file: Dict[str, FigureContext] = {}
    is_excel_source = False
    safe_source_id = str(state.get("source_id") or "").split(":")[-1] if state.get("source_id") else ""
    try:
        import asyncio
        import os
        import subprocess
        import tempfile

        from content_core.common import ProcessSourceState
        
        def _sync_extract(state, source_id):
            engine = state.get("document_engine")
            file_path = state.get("file_path")
            
            # Intercept MinerU extraction
            if engine == "mineru" and file_path and file_path.lower().endswith(('.pdf', '.ppt', '.pptx', '.doc', '.docx')):
                logger.info(f"Using MinerU to extract content from {file_path}")
                try:
                    with tempfile.TemporaryDirectory() as temp_dir:
                        if file_path.lower().endswith(('.ppt', '.pptx', '.doc', '.docx')):
                            from open_notebook.utils.office_converter import (
                                convert_to_modern_office_format,
                            )

                            file_path = convert_to_modern_office_format(file_path)
                            state["file_path"] = file_path

                        # Run mineru CLI
                        env = os.environ.copy()
                        if "HF_ENDPOINT" not in env:
                            env["HF_ENDPOINT"] = "https://hf-mirror.com"
                            
                        try:
                            # Enable table extraction enhancement
                            env["MINERU_TABLE_ENABLE"] = "true"
                            # Enable fast downloads
                            env["HF_HUB_ENABLE_HF_TRANSFER"] = "1"
                            # Use modelscope for faster downloads
                            env["MINERU_MODEL_SOURCE"] = "modelscope"
                            import sys
                            
                            logger.info("MinerU may need to download models on first run. Streaming output to console...")
                            subprocess.run([
                                "mineru",
                                "-p", file_path,
                                "-o", temp_dir,
                                "-m", "auto",
                                "--backend", "pipeline",
                            ], check=True, env=env, stdout=sys.stdout, stderr=sys.stderr)
                        except subprocess.CalledProcessError as e:
                            logger.error(f"MinerU extraction process failed with exit code {e.returncode}")
                            raise
                        
                        # Find the output directory (mineru creates a dir based on filename and model name)
                        base_name = os.path.splitext(os.path.basename(file_path))[0]
                        out_dir = os.path.join(temp_dir, base_name, "auto")
                        
                        target_dir = out_dir if os.path.exists(out_dir) else os.path.join(temp_dir, base_name)
                        
                        md_content = ""
                        if os.path.exists(target_dir):
                            for file in os.listdir(target_dir):
                                if file.endswith(".md"):
                                    with open(os.path.join(target_dir, file), "r", encoding="utf-8") as f:
                                        md_content = f.read()
                                    break
                            
                            # Copy images to persistent storage
                            import re
                            import shutil
                            
                            images_dir = os.path.join(target_dir, "images")
                            if os.path.exists(images_dir) and source_id:
                                safe_source_id = str(source_id).split(':')[-1]
                                project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
                                persistent_images_dir = os.path.join(project_root, "data", "uploads", "images", safe_source_id)
                                os.makedirs(persistent_images_dir, exist_ok=True)
                                
                                for img_file in os.listdir(images_dir):
                                    shutil.copy2(os.path.join(images_dir, img_file), os.path.join(persistent_images_dir, img_file))
                                
                                # Replace image paths in both Markdown and HTML formats
                                safe_images_url = f"/api/uploads/images/{safe_source_id}"
                                
                                # 1. Replace Markdown image syntax: ![alt](images/xxx.jpg)
                                md_content = re.sub(
                                    r'\!?\[.*?\]\((images/[^)]+)\)',
                                    lambda m: f"![](/api/uploads/images/{safe_source_id}/{os.path.basename(m.group(1))})",
                                    md_content
                                )
                                # 2. Replace HTML <img> tags: <img src="images/xxx.jpg" ...>
                                md_content = re.sub(
                                    r'<img\s+([^>]*?)src=["\']images/([^"\']+)["\']([^>]*)>',
                                    lambda m: f'<img {m.group(1)}src="{safe_images_url}/{os.path.basename(m.group(2))}"{m.group(3)}>',
                                    md_content
                                )
                                logger.debug(f"Replaced image paths pointing to {safe_images_url}")
                        
                        if md_content:
                            logger.info(f"Successfully extracted {len(md_content)} chars using MinerU.")
                            state["content"] = md_content
                            if not state.get("title") and "file_path" in state and state["file_path"]:
                                state["title"] = os.path.basename(state["file_path"])
                            # Bypass content_core's Pydantic validation which only allows 'auto', 'simple', 'docling'
                            state["document_engine"] = "auto"
                            return ProcessSourceState(**state)
                        else:
                            logger.warning("MinerU failed to produce markdown output. Falling back to simple engine.")
                            state["document_engine"] = "simple"
                except Exception as e:
                    logger.error(f"MinerU extraction failed: {e}. Falling back to simple engine.")
                    state["document_engine"] = "simple"
            elif engine == "mineru":
                logger.warning("MinerU does not support this file type. Falling back to simple engine.")
                state["document_engine"] = "simple"

            # Create a new event loop for this thread to run the async function
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            try:
                return loop.run_until_complete(extract_content(state))
            finally:
                loop.close()
                
        file_path = content_state.get("file_path", "")
        if should_bypass_content_core_for_image(file_path):
            content_state["content"] = content_state.get("content") or ""
            content_state["title"] = content_state.get("title") or os.path.basename(file_path)
            content_state["document_engine"] = "auto"
            processed_state = ProcessSourceState(**content_state)
            logger.info(
                f"Bypassing content-core extraction for standalone image source_id={state.get('source_id')} file_path={file_path}"
            )
        else:
            # 让 CPU 密集型任务在背景线程中运行，不阻塞主事件循环
            processed_state = await asyncio.to_thread(_sync_extract, content_state, state.get("source_id"))
        logger.info(f"Content extraction completed for source_id={state.get('source_id')}")

        file_ext = os.path.splitext(file_path)[1].lower() if file_path else ""
        excel_ext = file_ext
        is_excel_source = bool(file_path and file_ext in (".xls", ".xlsx", ".xlsm"))
        if file_path and file_ext in SUPPORTED_VISION_IMAGE_EXTENSIONS and safe_source_id:
            project_root = os.path.dirname(
                os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            )
            images_dir = os.path.join(
                project_root, "data", "uploads", "images", safe_source_id
            )
            os.makedirs(images_dir, exist_ok=True)
            image_filename = os.path.basename(file_path)
            target_path = os.path.join(images_dir, image_filename)
            if not os.path.exists(target_path):
                shutil.copy2(file_path, target_path)
            processed_state.content = processed_state.content or ""
            logger.info(
                f"Registered standalone image for vision description: {image_filename}"
            )
        if is_excel_source:
            processed_state.content = _trim_excel_empty_table_rows(
                _sanitize_excel_table_newlines(processed_state.content or "")
            )

            # Extract embedded images directly from .xlsx/.xlsm files
            if excel_ext == ".xls":
                logger.warning(
                    "Excel image extraction skipped for legacy .xls format. "
                    "Only .xlsx/.xlsm image extraction is supported."
                )
            else:
                try:
                    from openpyxl import load_workbook

                    source_id = state.get("source_id")
                    if source_id and file_path and safe_source_id:
                        project_root = os.path.dirname(
                            os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                        )
                        images_dir = os.path.join(
                            project_root, "data", "uploads", "images", safe_source_id
                        )
                        os.makedirs(images_dir, exist_ok=True)

                        existing_images = sorted(
                            [
                                f
                                for f in os.listdir(images_dir)
                                if f.lower().endswith(SUPPORTED_VISION_IMAGE_EXTENSIONS)
                            ]
                        )
                        if existing_images:
                            extracted_excel_figures = [
                                {"filename": name, "anchor": ""} for name in existing_images
                            ]
                            figure_contexts_by_file.update(
                                {
                                    name: FigureContext(
                                        filename=name,
                                        source_kind="excel",
                                        image_role="embedded_excel_image",
                                    )
                                    for name in existing_images
                                }
                            )
                            logger.info(
                                f"Reusing {len(existing_images)} existing Excel images for source_id={source_id}"
                            )
                        else:
                            workbook = load_workbook(file_path, data_only=True)
                            image_counter = 1
                            try:
                                for sheet in workbook.worksheets:
                                    images = getattr(sheet, "_images", []) or []
                                    sheet_rows = [
                                        [cell.value for cell in row]
                                        for row in sheet.iter_rows()
                                    ]
                                    for image in images:
                                        image_bytes = _extract_excel_image_bytes(image)
                                        if not image_bytes:
                                            continue

                                        ext = _infer_excel_image_ext(image)
                                        filename = f"excel_img_{image_counter:03d}.{ext}"
                                        image_path = os.path.join(images_dir, filename)
                                        with open(image_path, "wb") as f_img:
                                            f_img.write(image_bytes)

                                        anchor_label = _excel_anchor_label(image, sheet.title)
                                        anchor_cell = _excel_anchor_cell(anchor_label)

                                        extracted_excel_figures.append(
                                            {"filename": filename, "anchor": anchor_label}
                                        )
                                        figure_contexts_by_file[filename] = (
                                            _build_excel_context_from_anchor(
                                                sheet_rows,
                                                sheet.title,
                                                anchor_cell,
                                                filename,
                                                width=int(image.width) if image.width else None,
                                                height=int(image.height) if image.height else None,
                                            )
                                        )
                                        image_counter += 1
                            finally:
                                workbook.close()

                            if extracted_excel_figures:
                                logger.info(
                                    f"Extracted {len(extracted_excel_figures)} images from Excel workbook "
                                    f"for source_id={source_id}"
                                )
                            else:
                                logger.debug(
                                    f"No embedded images found in Excel workbook for source_id={source_id}"
                                )
                except Exception as img_e:
                    logger.warning(f"Excel image extraction failed (non-fatal): {img_e}")

        logger.debug(f"Extracted content length: {len(processed_state.content or '')} characters")
    except Exception as e:
        logger.error(f"Error during content extraction for source_id={state.get('source_id')}: {e}")
        raise

    # Describe extracted images with vision model
    try:
        source_id = state.get("source_id")
        if source_id:
            safe_source_id = str(source_id).split(':')[-1]
            project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            images_dir = os.path.join(project_root, "data", "uploads", "images", safe_source_id)

            if os.path.isdir(images_dir):
                image_files = sorted([
                    f for f in os.listdir(images_dir)
                    if f.lower().endswith(SUPPORTED_VISION_IMAGE_EXTENSIONS)
                ])
                if image_files:
                    from esperanto import LanguageModel

                    from open_notebook.ai.models import model_manager

                    defaults = await model_manager.get_defaults()
                    vision_model_id = defaults.default_vision_model
                    vision_model_kwargs: Dict[str, Any] = {}
                    if vision_model_id:
                        vision_model_config = await Model.get(vision_model_id)
                        vision_model_kwargs = _vision_model_inference_kwargs(
                            vision_model_config.provider
                        )
                    vision_model = await model_manager.get_vision_model(
                        **vision_model_kwargs
                    )
                    if vision_model and isinstance(vision_model, LanguageModel):
                        import base64

                        from langchain_core.messages import HumanMessage

                        vision_lc = vision_model.to_langchain()
                        mime_map = {
                            '.jpg': 'jpeg',
                            '.jpeg': 'jpeg',
                            '.png': 'png',
                            '.gif': 'gif',
                            '.webp': 'webp',
                            '.bmp': 'bmp',
                            '.tif': 'tiff',
                            '.tiff': 'tiff',
                            '.img': 'jpeg',
                        }
                        concurrency = _vision_concurrency()
                        semaphore = asyncio.Semaphore(concurrency)

                        async def _describe_one(img_file: str) -> tuple[str, str]:
                            """Describe a single image, returning (filename, description)."""
                            img_path = os.path.join(images_dir, img_file)
                            context = figure_contexts_by_file.get(img_file)
                            if context is None:
                                context = _default_figure_context(
                                    img_file,
                                    is_excel_source,
                                    safe_source_id,
                                )
                            try:
                                async with semaphore:
                                    with open(img_path, "rb") as f:
                                        img_data = base64.b64encode(f.read()).decode("utf-8")
                                    ext = os.path.splitext(img_file)[1].lower()
                                    mime_type = mime_map.get(ext, 'jpeg')
                                    vision_prompt = _build_vision_prompt(
                                        state.get("language"),
                                        context,
                                    )

                                    msg = HumanMessage(content=[
                                        {"type": "text", "text": vision_prompt},
                                        {"type": "image_url", "image_url": {"url": f"data:image/{mime_type};base64,{img_data}"}}
                                    ])

                                    response = await _invoke_vision_with_retries(
                                        lambda: vision_lc.ainvoke([msg])
                                    )
                                raw = response.content if hasattr(response, 'content') else str(response)
                                desc = _safe_figure_description(img_file, raw, context)
                                logger.info(f"Described image: {img_file}")
                                return (img_file, desc)
                            except Exception as img_e:
                                logger.warning(f"Failed to describe image {img_file}: {img_e}")
                                return (img_file, _vision_failure_description(img_e))

                        logger.info(
                            f"Starting concurrent image description for {len(image_files)} images "
                            f"(concurrency={concurrency})"
                        )
                        results = await asyncio.gather(*[_describe_one(f) for f in image_files])

                        descriptions: List[Dict[str, str]] = []
                        descriptions_by_file: Dict[str, str] = {}
                        for img_file, desc in results:
                            descriptions.append({"filename": img_file, "description": desc})
                            descriptions_by_file[img_file] = desc

                        logger.info(
                            f"Completed image descriptions: {len(descriptions)}/{len(image_files)}"
                        )

                        if is_excel_source:
                            figures = extracted_excel_figures or [
                                {"filename": name, "anchor": ""} for name in image_files
                            ]
                            if figures:
                                processed_state.content = (
                                    (processed_state.content or "")
                                    + _build_excel_figure_markdown(
                                        safe_source_id,
                                        figures,
                                        descriptions_by_file,
                                    )
                                )
                                logger.info(
                                    f"Added {len(figures)} extracted figure entries and descriptions "
                                    f"to source content"
                                )
                        elif descriptions:
                            processed_state.content = (
                                (processed_state.content or "")
                                + _build_standalone_image_markdown(
                                    safe_source_id,
                                    image_files,
                                    descriptions_by_file,
                                )
                            )
                            logger.info(f"Added {len(descriptions)} figure descriptions to source content")
                    else:
                        logger.debug("No vision model configured. Skipping image description.")
                        if is_excel_source:
                            figures = extracted_excel_figures or [
                                {"filename": name, "anchor": ""} for name in image_files
                            ]
                            if figures:
                                placeholder_descriptions = {
                                    fig["filename"]: "Vision model is not configured. Description unavailable."
                                    for fig in figures
                                }

                                processed_state.content = (
                                    (processed_state.content or "")
                                    + _build_excel_figure_markdown(
                                        safe_source_id,
                                        figures,
                                        placeholder_descriptions,
                                    )
                                )
                                logger.info(
                                    f"Added {len(figures)} extracted figure entries with placeholder descriptions "
                                    "to source content"
                                )
                        else:
                            placeholder_descriptions = {
                                filename: "Vision model is not configured. Description unavailable."
                                for filename in image_files
                            }
                            processed_state.content = (
                                (processed_state.content or "")
                                + _build_standalone_image_markdown(
                                    safe_source_id,
                                    image_files,
                                    placeholder_descriptions,
                                )
                            )
                            logger.info(
                                f"Added {len(image_files)} standalone image entries with placeholder descriptions "
                                "to source content"
                            )
    except Exception as e:
        logger.warning(f"Image description failed (non-fatal): {e}")

    if not processed_state.content or not processed_state.content.strip():
        url = processed_state.url or ""
        if url and ("youtube.com" in url or "youtu.be" in url):
            raise ValueError(
                "Could not extract content from this YouTube video. "
                "No transcript or subtitles are available. "
                "Try configuring a Speech-to-Text model in Settings "
                "to transcribe the audio instead."
            )
        raise ValueError(
            "Could not extract any text content from this source. "
            "The content may be empty, inaccessible, or in an unsupported format."
        )

    return {"content_state": processed_state}


async def save_source(state: SourceState) -> dict:
    content_state = state["content_state"]

    # Get existing source using the provided source_id
    source = await Source.get(state["source_id"])
    if not source:
        raise ValueError(f"Source with ID {state['source_id']} not found")

    # Update the source with processed content
    raw_original_filename = getattr(content_state, "original_filename", None)
    original_filename = raw_original_filename if isinstance(raw_original_filename, str) else None
    source.asset = Asset(
        url=content_state.url,
        file_path=content_state.file_path,
        original_filename=original_filename,
    )
    source.full_text = content_state.content

    # Preserve user-set title; only overwrite placeholder or empty titles
    if content_state.title and (not source.title or source.title == "Processing..."):
        source.title = content_state.title

    await source.save()

    # NOTE: Notebook associations are created by the API immediately for UI responsiveness
    # No need to create them here to avoid duplicate edges

    if state["embed"]:
        if source.full_text and source.full_text.strip():
            logger.debug("Embedding content for vector search")
            await source.vectorize()
        else:
            logger.warning(
                f"Source {source.id} has no text content to embed, skipping vectorization"
            )

    return {"source": source}


def trigger_transformations(state: SourceState, config: RunnableConfig) -> List[Send]:
    if len(state["apply_transformations"]) == 0:
        return []

    to_apply = state["apply_transformations"]
    logger.debug(f"Applying transformations {to_apply}")

    return [
        Send(
            "transform_content",
            {
                "source": state["source"],
                "transformation": t,
            },
        )
        for t in to_apply
    ]


async def transform_content(state: TransformationState) -> Optional[dict]:
    source = state["source"]
    content = source.full_text
    if not content:
        return None
    transformation: Transformation = state["transformation"]

    logger.info(f"Submitting background job for transformation {transformation.title or transformation.name}")
    from surreal_commands import submit_command
    submit_command(
        "open_notebook",
        "run_transformation",
        {
            "source_id": str(source.id),
            "transformation_id": str(transformation.id)
        }
    )
    
    return {
        "transformation": [
            {
                "output": "Transformation job submitted to background worker",
                "transformation_name": transformation.title or transformation.name,
            }
        ]
    }


# Create and compile the workflow
workflow = StateGraph(SourceState)

# Add nodes
workflow.add_node("content_process", content_process)
workflow.add_node("save_source", save_source)
workflow.add_node("transform_content", transform_content)
# Define the graph edges
workflow.add_edge(START, "content_process")
workflow.add_edge("content_process", "save_source")
workflow.add_conditional_edges(
    "save_source", trigger_transformations, ["transform_content"]
)
workflow.add_edge("transform_content", END)

# Compile the graph
source_graph = workflow.compile()
